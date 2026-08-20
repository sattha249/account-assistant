import datetime
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils import normalize_category_name, category_key, is_same_category


def prepare_target_sheet(wb: openpyxl.Workbook, base_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    If sheet with base_name exists in workbook, deletes/replaces it directly.
    Returns new worksheet created with title base_name.
    """
    if base_name in wb.sheetnames:
        del wb[base_name]
    return wb.create_sheet(title=base_name)


def is_orange_fill(cell) -> bool:
    """Returns True if the cell has Orange/Amber background fill."""
    if cell and cell.fill and cell.fill.fill_type and cell.fill.start_color and cell.fill.start_color.rgb:
        rgb_str = str(cell.fill.start_color.rgb).upper()
        if "FFC000" in rgb_str or "FFD700" in rgb_str or "ORANGE" in rgb_str:
            return True
    return False


def get_skipped_orange_rows(ws_sales: openpyxl.worksheet.worksheet.Worksheet, max_r: int) -> set:
    """
    Scans sheet 'ขาย' for rows with Orange fill.
    For each orange row r, checks row r-1 and row r+1.
    If Date(r) == Date(adj) AND Qty(r) == Qty(adj), marks row r to be skipped.
    """
    skipped_rows = set()
    
    for r in range(2, max_r + 1):
        cell_b = ws_sales.cell(row=r, column=2)
        cell_d = ws_sales.cell(row=r, column=4)
        cell_e = ws_sales.cell(row=r, column=5)
        
        is_orange = any(is_orange_fill(c) for c in (cell_b, cell_d, cell_e))
        
        if is_orange:
            d_curr = ws_sales.cell(row=r, column=2).value
            q_curr = ws_sales.cell(row=r, column=4).value
            
            # Compare with prev row (r-1)
            match_prev = False
            if r > 2:
                d_prev = ws_sales.cell(row=r - 1, column=2).value
                q_prev = ws_sales.cell(row=r - 1, column=4).value
                if d_curr == d_prev and q_curr == q_prev and q_curr is not None:
                    match_prev = True
            
            # Compare with next row (r+1)
            match_next = False
            if r < max_r:
                d_next = ws_sales.cell(row=r + 1, column=2).value
                q_next = ws_sales.cell(row=r + 1, column=4).value
                if d_curr == d_next and q_curr == q_next and q_curr is not None:
                    match_next = True
            
            if match_prev or match_next:
                skipped_rows.add(r)
                
    return skipped_rows


def parse_barn_string(barn_str, row_qty=None) -> list:
    """
    Parses barn strings such as 'T6=2 , T4=1 , T5=1', 'N1 , T6', '1F1=12 , 1F2=14', 'T13=1 รวมรก', 'T8=1 , N3=1 N1=3'.
    Extracts all barn quantity patterns anywhere in the string regardless of missing commas.
    Returns list of tuples: [(sub_barn_id, quantity), ...]
    """
    if not barn_str:
        return []
    matches = re.findall(r'([A-Za-z0-9]+)(?:\s*=\s*(\d+))?', str(barn_str))
    parsed = []
    ignore_words = {'รวม', 'รวมรก', 'ขาย', 'ซาก', 'บาท', 'ตัว', 'KG', 'กิโล', 'เดือน', 'วัน'}
    for sub_b, qty_str in matches:
        sub_upper = sub_b.upper()
        if sub_upper in ignore_words:
            continue
        if qty_str:
            parsed.append((sub_upper, int(qty_str)))
        else:
            qty = int(row_qty) if (row_qty and isinstance(row_qty, (int, float)) and len(matches) == 1) else 1
            parsed.append((sub_upper, qty))
    return parsed


def get_main_barn_group(sub_barn: str) -> str:
    """
    Classifies a sub-barn ID into one of 3 main barn groups:
    - 'กลุ่มเล้า T': Sub-barn starts with 'T' followed by digits (e.g. T1..T16)
    - 'กลุ่มเล้า N': Sub-barn starts with 'N' followed by digits (e.g. N1..N8)
    - 'กลุ่มเล้าอื่นๆ': Anything else (e.g. 1F1, 2F1, 1M5, 2M6, G1, Q, AI...)
    """
    sub = sub_barn.upper().strip()
    if sub.startswith("T") and sub[1:].isdigit():
        return "กลุ่มเล้า T"
    elif sub.startswith("N") and sub[1:].isdigit():
        return "กลุ่มเล้า N"
    else:
        return "กลุ่มเล้าอื่นๆ"


def barn_sort_key(sub_barn: str) -> tuple:
    """
    Generates a natural sort key for sub-barns (e.g. T1, T2..T16, N1..N8, 1F1..2M6).
    """
    sub = sub_barn.upper().strip()
    m = re.search(r'(\d+)', sub)
    num = int(m.group(1)) if m else 0
    alpha = re.sub(r'\d+', '', sub)
    return (alpha, num, sub)


def detect_weeks_and_categories(ws_sales: openpyxl.worksheet.worksheet.Worksheet) -> tuple:
    """
    Detects week blocks (WK.xx) and date ranges in sheet 'ขาย'.
    Also extracts all distinct product categories across the sheet.
    
    Returns:
        (weeks: list of dicts, categories: list of dicts)
    """
    max_r = ws_sales.max_row
    weeks = []
    start_r = 2

    # Step 1: Scan for week separators (Column B starting with 'WK.')
    for r in range(2, max_r + 1):
        cell_b = ws_sales.cell(row=r, column=2).value
        cell_c = ws_sales.cell(row=r, column=3).value

        if cell_b and str(cell_b).strip().startswith("WK."):
            d_str = str(cell_c).strip() if cell_c else ""
            if d_str.startswith("วันที่ "):
                d_str = d_str.replace("วันที่ ", "")

            weeks.append({
                "name": str(cell_b).strip(),
                "date_range": d_str,
                "start_row": start_r,
                "end_row": r
            })
            start_r = r + 1

    # If remaining rows exist after the last WK separator
    if start_r <= max_r:
        dates = []
        for r in range(start_r, max_r + 1):
            v = ws_sales.cell(row=r, column=2).value
            if hasattr(v, "strftime"):
                y = v.year if v.year < 2500 else v.year - 543
                dates.append(f"{v.day}/{v.month}/{str(y)[-2:]}")
        
        d_range = f"{dates[0]} - {dates[-1]}" if len(dates) > 1 else (dates[0] if dates else "ท้ายเดือน")
        
        has_content = any(ws_sales.cell(row=r, column=5).value for r in range(start_r, max_r + 1))
        if has_content:
            wk_num = len(weeks) + 32 if weeks else 1
            weeks.append({
                "name": f"WK.{wk_num}" if weeks else "สรุปทั้งเดือน",
                "date_range": d_range,
                "start_row": start_r,
                "end_row": max_r
            })

    if not weeks:
        weeks = [{
            "name": "สรุปทั้งเดือน",
            "date_range": "ทั้งหมด",
            "start_row": 2,
            "end_row": max_r
        }]

    # Step 2: Calculate date ranges for weeks if empty
    for wk in weeks:
        if not wk["date_range"]:
            dates = []
            for r in range(wk["start_row"], wk["end_row"] + 1):
                v = ws_sales.cell(row=r, column=2).value
                if hasattr(v, "strftime"):
                    y = v.year if v.year < 2500 else v.year - 543
                    dates.append(f"{v.day}/{v.month}/{str(y)[-2:]}")
            if dates:
                wk["date_range"] = f"{dates[0]} - {dates[-1]}" if len(dates) > 1 else dates[0]
            else:
                wk["date_range"] = "-"

    # Step 3: Extract distinct categories from Column E (Col 5)
    categories = []
    seen_keys = set()

    for r in range(2, max_r + 1):
        c_val = ws_sales.cell(row=r, column=5).value # Column E: ประเภท
        clean_name = normalize_category_name(c_val)
        c_key = category_key(clean_name)

        if clean_name and c_key not in seen_keys and c_key not in ("รวม", "ประเภท", "วันที่", "ชื่อลูกค้า"):
            seen_keys.add(c_key)
            categories.append({
                "name": clean_name,
                "key": c_key
            })

    return weeks, categories


def add_summary_sheet(file_path: str) -> dict:
    """
    Processes the specified Excel file, creates/replaces a styled multi-column weekly matrix 'สรุปยอด' sheet,
    linking formulas dynamically per week block and date range. (Menu 1)
    """
    if not os.path.exists(file_path):
        return {
            "success": False,
            "error": f"ไม่พบไฟล์: {file_path}",
            "sheet_name": None,
        }

    try:
        wb = openpyxl.load_workbook(file_path)

        if "ขาย" not in wb.sheetnames:
            return {
                "success": False,
                "error": "ไม่พบชีท 'ขาย' ในไฟล์ Excel ที่ระบุ",
                "sheet_name": None,
            }

        ws_sales = wb["ขาย"]
        max_r = ws_sales.max_row

        # Detect skipped orange duplicate rows
        skipped_orange_rows = get_skipped_orange_rows(ws_sales, max_r)

        # Detect weeks and categories dynamically
        weeks, categories = detect_weeks_and_categories(ws_sales)

        if not categories:
            return {
                "success": False,
                "error": "ไม่พบประเภทสินค้าในคอลัมน์ E ของชีท 'ขาย'",
                "sheet_name": None
            }

        # Replaces existing 'สรุปยอด' sheet if present
        target_sheet_name = "สรุปยอด"
        ws_summary = prepare_target_sheet(wb, target_sheet_name)

        # Styles
        font_header_1 = Font(name="Calibri", size=11, bold=True, color="000000")
        font_header_2 = Font(name="Calibri", size=10, bold=False, italic=True, color="333333")
        font_data = Font(name="Calibri", size=11, bold=False, color="000000")
        font_total = Font(name="Calibri", size=11, bold=True, color="000000")

        fill_header_1 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        fill_header_2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        thin_side = Side(border_style="thin", color="D9D9D9")
        thick_side = Side(border_style="medium", color="000000")
        double_side = Side(border_style="double", color="000000")

        border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_total = Border(left=thin_side, right=thin_side, top=thin_side, bottom=double_side)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        num_format = "#,##0"

        # --- Row 1: Header Row 1 (Week Names) ---
        c_item_1 = ws_summary.cell(row=1, column=1, value="รายการ")
        c_item_1.font = font_header_1
        c_item_1.fill = fill_header_1
        c_item_1.alignment = align_center
        c_item_1.border = Border(left=thick_side, right=thin_side, top=thick_side, bottom=thin_side)

        col_idx = 2
        for wk in weeks:
            cell = ws_summary.cell(row=1, column=col_idx, value=wk["name"])
            cell.font = font_header_1
            cell.fill = fill_header_1
            cell.alignment = align_center
            cell.border = Border(left=thin_side, right=thin_side, top=thick_side, bottom=thin_side)
            col_idx += 1

        cell_total_hdr1 = ws_summary.cell(row=1, column=col_idx, value="รวมทั้งเดือน")
        cell_total_hdr1.font = font_header_1
        cell_total_hdr1.fill = fill_header_1
        cell_total_hdr1.alignment = align_center
        cell_total_hdr1.border = Border(left=thin_side, right=thick_side, top=thick_side, bottom=thin_side)

        # --- Row 2: Header Row 2 (Date Ranges) ---
        c_item_2 = ws_summary.cell(row=2, column=1, value="ช่วงวันที่")
        c_item_2.font = font_header_2
        c_item_2.fill = fill_header_2
        c_item_2.alignment = align_center
        c_item_2.border = Border(left=thick_side, right=thin_side, top=thin_side, bottom=thick_side)

        col_idx = 2
        for wk in weeks:
            cell = ws_summary.cell(row=2, column=col_idx, value=wk["date_range"])
            cell.font = font_header_2
            cell.fill = fill_header_2
            cell.alignment = align_center
            cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_side)
            col_idx += 1

        cell_total_hdr2 = ws_summary.cell(row=2, column=col_idx, value="")
        cell_total_hdr2.font = font_header_2
        cell_total_hdr2.fill = fill_header_2
        cell_total_hdr2.alignment = align_center
        cell_total_hdr2.border = Border(left=thin_side, right=thick_side, top=thin_side, bottom=thick_side)

        max_name_len = 10
        total_pigs_qty = 0

        # --- Data Rows (Rows 3..M+2) ---
        current_row = 3
        for cat in categories:
            c_name = ws_summary.cell(row=current_row, column=1, value=cat["name"])
            c_name.font = font_data
            c_name.alignment = align_left
            c_name.border = border_cell

            max_name_len = max(max_name_len, len(cat["name"]))

            col_idx = 2
            for wk in weeks:
                start_r = wk["start_row"]
                end_r = wk["end_row"]
                
                valid_cell_refs = []
                for r in range(start_r, end_r + 1):
                    if r in skipped_orange_rows:
                        continue
                    c_val = ws_sales.cell(row=r, column=5).value
                    q_val = ws_sales.cell(row=r, column=4).value
                    clean_c = normalize_category_name(c_val)
                    if is_same_category(clean_c, cat["name"]):
                        valid_cell_refs.append(f"'ขาย'!D{r}")
                        if isinstance(q_val, (int, float)):
                            total_pigs_qty += int(q_val)

                c_val = ws_summary.cell(row=current_row, column=col_idx)
                c_val.font = font_data
                c_val.alignment = align_right
                c_val.number_format = num_format
                c_val.border = border_cell

                if not valid_cell_refs:
                    c_val.value = 0
                elif len(valid_cell_refs) == 1:
                    c_val.value = f"={valid_cell_refs[0]}"
                else:
                    c_val.value = f"=SUM({', '.join(valid_cell_refs)})"

                col_idx += 1

            last_wk_col_letter = get_column_letter(col_idx - 1)
            c_row_total = ws_summary.cell(row=current_row, column=col_idx)
            c_row_total.font = font_total
            c_row_total.alignment = align_right
            c_row_total.number_format = num_format
            c_row_total.border = border_cell
            c_row_total.value = f"=SUM(B{current_row}:{last_wk_col_letter}{current_row})"

            current_row += 1

        # --- Bottom Total Row (รวม) ---
        c_tot_name = ws_summary.cell(row=current_row, column=1, value="รวม")
        c_tot_name.font = font_total
        c_tot_name.alignment = align_center
        c_tot_name.border = border_total

        last_data_row = current_row - 1
        col_idx = 2
        for wk in weeks:
            col_letter = get_column_letter(col_idx)
            c_col_tot = ws_summary.cell(row=current_row, column=col_idx)
            c_col_tot.font = font_total
            c_col_tot.alignment = align_right
            c_col_tot.number_format = num_format
            c_col_tot.border = border_total
            c_col_tot.value = f"=SUM({col_letter}3:{col_letter}{last_data_row})"
            col_idx += 1

        total_col_letter = get_column_letter(col_idx)
        c_grand_tot = ws_summary.cell(row=current_row, column=col_idx)
        c_grand_tot.font = font_total
        c_grand_tot.alignment = align_right
        c_grand_tot.number_format = num_format
        c_grand_tot.border = border_total
        c_grand_tot.value = f"=SUM({total_col_letter}3:{total_col_letter}{last_data_row})"

        # Column Widths
        ws_summary.column_dimensions['A'].width = max(24, max_name_len + 6)
        for c in range(2, col_idx + 1):
            c_let = get_column_letter(c)
            ws_summary.column_dimensions[c_let].width = 18

        # Save workbook
        wb.save(file_path)

        summary_lines = [
            f"• สัปดาห์ทั้งหมด: {len(weeks)} สัปดาห์",
            f"• ประเภทสุกร/สินค้า: {len(categories)} รายการ",
            f"• ยอดขายรวมทั้งสิ้น: {total_pigs_qty:,} ตัว",
            f"• บรรทัดสีส้มซ้ำที่ข้าม: {len(skipped_orange_rows)} บรรทัด"
        ]
        summary_text = "\n".join(summary_lines)

        return {
            "success": True,
            "sheet_name": target_sheet_name,
            "file_path": file_path,
            "weeks_count": len(weeks),
            "categories_count": len(categories),
            "total_qty": total_pigs_qty,
            "skipped_orange_count": len(skipped_orange_rows),
            "summary_text": summary_text,
            "error": None
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"เกิดข้อผิดพลาดในการสร้างชีทสรุปยอด: {str(e)}",
            "sheet_name": None
        }


def add_barn_summary_sheet(file_path: str) -> dict:
    """
    Processes the specified Excel file and creates/replaces a styled 'สรุปยอดแยกเล้า' sheet (Menu 2).
    Generates a Weekly Matrix breakdown where EACH week contains 3 Main Barn Tables:
    1. กลุ่มเล้า T (T1..T16)
    2. กลุ่มเล้า N (N1..N8)
    3. กลุ่มเล้าอื่นๆ (1F1, 2F1, 1M5, 2M6, G1, Q, AI...)
    
    AT THE END OF EACH WEEK: Inserts ONE SINGLE SUMMARY ROW ('รวมทุกเล้าประจำสัปดาห์') 
    summing up all barns in that week, broken down by category!
    Includes Orange Fill Duplicate Row Exclusion Rule.
    Includes all categories in every table, defaulting to 0 for empty cells.
    Handles Merged Cells in Column N seamlessly.
    """
    if not os.path.exists(file_path):
        return {
            "success": False,
            "error": f"ไม่พบไฟล์: {file_path}",
            "sheet_name": None,
        }

    try:
        wb = openpyxl.load_workbook(file_path)

        if "ขาย" not in wb.sheetnames:
            return {
                "success": False,
                "error": "ไม่พบชีท 'ขาย' ในไฟล์ Excel ที่ระบุ",
                "sheet_name": None,
            }

        ws_sales = wb["ขาย"]
        max_r = ws_sales.max_row

        # Build merged cell lookup map for Column 14 (เล้า)
        merged_n_map = {}
        for rng in ws_sales.merged_cells.ranges:
            if rng.min_col <= 14 <= rng.max_col:
                top_val = ws_sales.cell(row=rng.min_row, column=14).value
                for r in range(rng.min_row, rng.max_row + 1):
                    merged_n_map[r] = (top_val, rng.min_row)

        # Detect skipped orange duplicate rows
        skipped_orange_rows = get_skipped_orange_rows(ws_sales, max_r)

        # Detect weeks and all categories dynamically
        weeks, categories = detect_weeks_and_categories(ws_sales)

        if not categories:
            return {
                "success": False,
                "error": "ไม่พบประเภทสินค้าในคอลัมน์ E ของชีท 'ขาย'",
                "sheet_name": None
            }

        # Structure: weekly_data[wk_name][main_group][sub_barn][category_key] = count
        weekly_data = {}
        all_t_barns = set()
        all_n_barns = set()
        all_other_barns = set()
        total_pigs_qty = 0

        for wk in weeks:
            weekly_data[wk["name"]] = {}
            for r in range(wk["start_row"], wk["end_row"] + 1):
                if r in skipped_orange_rows:
                    continue

                # Handle Merged Cells in Column N
                if r in merged_n_map:
                    barn_str, min_r = merged_n_map[r]
                    if r != min_r:
                        continue
                else:
                    barn_str = ws_sales.cell(row=r, column=14).value

                c_val = ws_sales.cell(row=r, column=5).value     # Column E: ประเภท
                q_val = ws_sales.cell(row=r, column=4).value     # Column D: จำนวนตัว

                clean_cat = normalize_category_name(c_val)
                c_key = category_key(clean_cat)

                if not clean_cat or c_key in ("รวม", "ประเภท", "วันที่", "ชื่อลูกค้า"):
                    continue

                parsed_barns = parse_barn_string(barn_str, q_val)
                for sub_b, b_qty in parsed_barns:
                    main_grp = get_main_barn_group(sub_b)
                    if main_grp == "กลุ่มเล้า T":
                        all_t_barns.add(sub_b)
                    elif main_grp == "กลุ่มเล้า N":
                        all_n_barns.add(sub_b)
                    else:
                        all_other_barns.add(sub_b)

                    total_pigs_qty += b_qty

                    if main_grp not in weekly_data[wk["name"]]:
                        weekly_data[wk["name"]][main_grp] = {}
                    if sub_b not in weekly_data[wk["name"]][main_grp]:
                        weekly_data[wk["name"]][main_grp][sub_b] = {}
                    if c_key not in weekly_data[wk["name"]][main_grp][sub_b]:
                        weekly_data[wk["name"]][main_grp][sub_b][c_key] = 0
                    weekly_data[wk["name"]][main_grp][sub_b][c_key] += b_qty

        # Replaces existing 'สรุปยอดแยกเล้า' sheet if present
        target_sheet_name = "สรุปยอดแยกเล้า"
        ws_barn = prepare_target_sheet(wb, target_sheet_name)

        # Styles
        font_wk_banner = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        font_banner = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_hdr = Font(name="Calibri", size=11, bold=True, color="000000")
        font_data = Font(name="Calibri", size=11, bold=False, color="000000")
        font_tot = Font(name="Calibri", size=11, bold=True, color="000000")
        font_wk_tot = Font(name="Calibri", size=11, bold=True, color="000000")

        fill_wk_banner = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy Week Banner
        fill_banner = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")    # Medium Blue Barn Group Banner
        fill_hdr = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")       # Light Blue Table Header
        fill_zebra = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")     # Soft Zebra
        fill_wk_tot = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")    # Warm Gold Fill for Single Week Summary Row

        thin_side = Side(border_style="thin", color="D9D9D9")
        thick_side = Side(border_style="medium", color="000000")
        double_side = Side(border_style="double", color="000000")

        border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_tot = Border(left=thin_side, right=thin_side, top=thin_side, bottom=double_side)
        border_wk_tot = Border(left=thick_side, right=thick_side, top=thick_side, bottom=double_side)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        num_format = "#,##0"
        main_groups_order = ["กลุ่มเล้า T", "กลุ่มเล้า N", "กลุ่มเล้าอื่นๆ"]
        tot_table_cols = 1 + len(categories) + 1 # SubBarn + Categories + Total

        current_row = 1
        max_col_width_a = 14

        for wk in weeks:
            wk_name = wk["name"]
            date_range = wk["date_range"]
            week_group_total_rows = [] # Track total row indices of T, N, Other barn tables for this week
            
            # 1. Week Header Banner Row
            ws_barn.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=tot_table_cols)
            wk_banner_cell = ws_barn.cell(row=current_row, column=1, value=f"[ {wk_name}  (ช่วงวันที่ {date_range}) ]")
            wk_banner_cell.font = font_wk_banner
            wk_banner_cell.fill = fill_wk_banner
            wk_banner_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_barn.row_dimensions[current_row].height = 26
            current_row += 1

            for main_grp in main_groups_order:
                if main_grp not in weekly_data[wk_name]:
                    continue

                sub_barns = sorted(weekly_data[wk_name][main_grp].keys(), key=barn_sort_key)
                if not sub_barns:
                    continue

                # 2. Main Group Banner Row
                ws_barn.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=tot_table_cols)
                banner_cell = ws_barn.cell(row=current_row, column=1, value=f"  [ {main_grp} ]")
                banner_cell.font = font_banner
                banner_cell.fill = fill_banner
                banner_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                ws_barn.row_dimensions[current_row].height = 22
                current_row += 1

                # 3. Table Header Row
                c_hdr_sub = ws_barn.cell(row=current_row, column=1, value="เล้า")
                c_hdr_sub.font = font_hdr
                c_hdr_sub.fill = fill_hdr
                c_hdr_sub.alignment = align_center
                c_hdr_sub.border = Border(left=thick_side, right=thin_side, top=thick_side, bottom=thick_side)

                c_col = 2
                for cat in categories:
                    cell = ws_barn.cell(row=current_row, column=c_col, value=cat["name"])
                    cell.font = font_hdr
                    cell.fill = fill_hdr
                    cell.alignment = align_center
                    cell.border = Border(left=thin_side, right=thin_side, top=thick_side, bottom=thick_side)
                    c_col += 1

                c_hdr_tot = ws_barn.cell(row=current_row, column=c_col, value="รวม")
                c_hdr_tot.font = font_hdr
                c_hdr_tot.fill = fill_hdr
                c_hdr_tot.alignment = align_center
                c_hdr_tot.border = Border(left=thin_side, right=thick_side, top=thick_side, bottom=thick_side)

                current_row += 1
                table_data_start_row = current_row

                # 4. Sub-Barn Data Rows
                for idx, sb in enumerate(sub_barns):
                    c_sb = ws_barn.cell(row=current_row, column=1, value=sb)
                    c_sb.font = font_hdr
                    c_sb.alignment = align_center
                    c_sb.border = border_cell

                    max_col_width_a = max(max_col_width_a, len(sb) + 6)
                    is_even = (idx % 2 == 0)

                    c_col = 2
                    for cat in categories:
                        qty = weekly_data[wk_name][main_grp][sb].get(cat["key"], 0)
                        cell = ws_barn.cell(row=current_row, column=c_col, value=qty)
                        cell.font = font_data
                        cell.alignment = align_right
                        cell.number_format = num_format
                        cell.border = border_cell
                        if is_even:
                            cell.fill = fill_zebra
                        c_col += 1

                    # Horizontal Row Total Formula
                    last_cat_letter = get_column_letter(c_col - 1)
                    c_row_tot = ws_barn.cell(row=current_row, column=c_col)
                    c_row_tot.font = font_tot
                    c_row_tot.alignment = align_right
                    c_row_tot.number_format = num_format
                    c_row_tot.border = border_cell
                    c_row_tot.value = f"=SUM(B{current_row}:{last_cat_letter}{current_row})"
                    if is_even:
                        c_row_tot.fill = fill_zebra

                    current_row += 1

                table_data_end_row = current_row - 1

                # 5. Table Bottom Total Row
                c_grp_tot_label = ws_barn.cell(row=current_row, column=1, value="รวม")
                c_grp_tot_label.font = font_tot
                c_grp_tot_label.alignment = align_center
                c_grp_tot_label.border = border_tot

                c_col = 2
                for cat in categories:
                    col_let = get_column_letter(c_col)
                    c_tot_cell = ws_barn.cell(row=current_row, column=c_col)
                    c_tot_cell.font = font_tot
                    c_tot_cell.alignment = align_right
                    c_tot_cell.number_format = num_format
                    c_tot_cell.border = border_tot
                    c_tot_cell.value = f"=SUM({col_let}{table_data_start_row}:{col_let}{table_data_end_row})"
                    c_col += 1

                # Group Grand Total
                last_cat_letter = get_column_letter(c_col - 1)
                total_col_letter = get_column_letter(c_col)
                c_grp_grand = ws_barn.cell(row=current_row, column=c_col)
                c_grp_grand.font = font_tot
                c_grp_grand.alignment = align_right
                c_grp_grand.number_format = num_format
                c_grp_grand.border = border_tot
                c_grp_grand.value = f"=SUM({total_col_letter}{table_data_start_row}:{total_col_letter}{table_data_end_row})"

                # Store total row index for week summary
                week_group_total_rows.append(current_row)

                # Spacer row between barn group tables
                current_row += 2

            # --- SINGLE SUMMARY ROW AT THE END OF THE WEEK (รวมทุกเล้าประจำสัปดาห์) ---
            if week_group_total_rows:
                c_wk_tot_label = ws_barn.cell(row=current_row, column=1, value=f"รวมทุกเล้า ({wk_name})")
                c_wk_tot_label.font = font_wk_tot
                c_wk_tot_label.fill = fill_wk_tot
                c_wk_tot_label.alignment = align_center
                c_wk_tot_label.border = border_wk_tot

                c_col = 2
                for cat in categories:
                    col_let = get_column_letter(c_col)
                    refs = [f"{col_let}{r_idx}" for r_idx in week_group_total_rows]
                    
                    c_wk_cell = ws_barn.cell(row=current_row, column=c_col)
                    c_wk_cell.font = font_wk_tot
                    c_wk_cell.fill = fill_wk_tot
                    c_wk_cell.alignment = align_right
                    c_wk_cell.number_format = num_format
                    c_wk_cell.border = border_wk_tot
                    c_wk_cell.value = f"=SUM({', '.join(refs)})" if len(refs) > 1 else f"={refs[0]}"
                    c_col += 1

                # Week Grand Total Formula (Sum across categories for this week)
                last_cat_letter = get_column_letter(c_col - 1)
                c_wk_grand = ws_barn.cell(row=current_row, column=c_col)
                c_wk_grand.font = font_wk_tot
                c_wk_grand.fill = fill_wk_tot
                c_wk_grand.alignment = align_right
                c_wk_grand.number_format = num_format
                c_wk_grand.border = border_wk_tot
                c_wk_grand.value = f"=SUM(B{current_row}:{last_cat_letter}{current_row})"

                current_row += 1

            # Spacer rows between weeks
            current_row += 2

        # Column Widths
        ws_barn.column_dimensions['A'].width = max(22, max_col_width_a)
        for c in range(2, tot_table_cols + 1):
            c_let = get_column_letter(c)
            ws_barn.column_dimensions[c_let].width = 16

        wb.save(file_path)

        total_sub_barns = len(all_t_barns) + len(all_n_barns) + len(all_other_barns)
        summary_lines = [
            f"• สัปดาห์ทั้งหมด: {len(weeks)} สัปดาห์",
            f"• เล้าย่อยทั้งหมด: {total_sub_barns} เล้า (กลุ่ม T: {len(all_t_barns)}, กลุ่ม N: {len(all_n_barns)}, อื่นๆ: {len(all_other_barns)})",
            f"• ยอดขายรวมทั้งสิ้น: {total_pigs_qty:,} ตัว",
            f"• บรรทัดสีส้มซ้ำที่ข้าม: {len(skipped_orange_rows)} บรรทัด"
        ]
        summary_text = "\n".join(summary_lines)

        return {
            "success": True,
            "sheet_name": target_sheet_name,
            "file_path": file_path,
            "weeks_count": len(weeks),
            "total_sub_barns": total_sub_barns,
            "t_barns_count": len(all_t_barns),
            "n_barns_count": len(all_n_barns),
            "other_barns_count": len(all_other_barns),
            "total_qty": total_pigs_qty,
            "skipped_orange_count": len(skipped_orange_rows),
            "summary_text": summary_text,
            "error": None
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"เกิดข้อผิดพลาดในการสร้างชีทสรุปยอดแยกเล้า: {str(e)}",
            "sheet_name": None
        }
